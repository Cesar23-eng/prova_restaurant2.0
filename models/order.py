# models/order.py
import os
import datetime
from collections import defaultdict
from typing import Dict, List, Tuple, Optional


def _get_data_dir() -> str:
    """Carpeta data/YYYY-MM-DD/ junto al ejecutable o al proyecto."""
    import sys
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(sys.argv[0]))
    today = datetime.date.today().strftime("%Y-%m-%d")
    path = os.path.join(base, "data", today)
    os.makedirs(path, exist_ok=True)
    return path


def _get_sequence_file() -> str:
    import sys
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(sys.argv[0]))
    data_dir = os.path.join(base, "data")
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, "sequence.txt")


def _next_order_number() -> str:
    """Lee, incrementa y guarda el contador de pedidos. Devuelve '#0001'."""
    seq_file = _get_sequence_file()
    try:
        with open(seq_file, "r") as f:
            n = int(f.read().strip())
    except (FileNotFoundError, ValueError):
        n = 0
    n += 1
    with open(seq_file, "w") as f:
        f.write(str(n))
    return f"#{n:04d}"


def _write_audit(action: str, table_name: str, detail: str = ""):
    """Agrega una linea al log de auditoria del dia."""
    data_dir = _get_data_dir()
    today = datetime.date.today().strftime("%Y-%m-%d")
    log_file = os.path.join(data_dir, f"audit_log_{today}.txt")
    now = datetime.datetime.now().strftime("%H:%M:%S")
    line = f"[{now}] {action:12s} | {table_name}"
    if detail:
        line += f" | {detail}"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(line + "\n")


class OrderManager:
    def __init__(self):
        self.table_orders = defaultdict(list)
        self.payment_status = {}
        self.payment_details = {}
        self.delivery_details = {}
        self.order_numbers = {}      # table_name -> '#0001'
        self.current_table = None
        self.order_type = "En el local"

    # -----------------------------------------------
    #  Utilidades de nombre
    # -----------------------------------------------
    @staticmethod
    def _norm(name: str) -> str:
        return " ".join(name.split()).strip().lower()

    def name_exists(self, name: str) -> bool:
        n = self._norm(name)
        return any(self._norm(k) == n for k in self.table_orders.keys())

    def suggest_name(self, base: str) -> str:
        base = " ".join(base.split()).strip()
        if not self.name_exists(base):
            return base
        i = 2
        while True:
            candidate = f"{base} {i}"
            if not self.name_exists(candidate):
                return candidate
            i += 1

    # -----------------------------------------------
    #  Gestion de mesas / ordenes
    # -----------------------------------------------
    def create_table(self, name: str) -> Tuple[bool, str]:
        base = " ".join((name or "").split()).strip()
        if not base:
            return False, "Nombre invalido"
        if self.name_exists(base):
            return False, self.suggest_name(base)
        self.table_orders[base] = []
        self.payment_status.pop(base, None)
        self.payment_details.pop(base, None)
        self.delivery_details.pop(base, None)
        num = _next_order_number()
        self.order_numbers[base] = num
        _write_audit("CREAR", base, f"numero={num}")
        return True, base

    def get_order_number(self, table_name: str) -> str:
        return self.order_numbers.get(table_name, "")

    def set_current_table(self, table_name: str):
        self.current_table = table_name

    def add_item_to_order(self, category, dish, variant, price):
        if not self.current_table:
            raise ValueError("No hay mesa seleccionada")
        # Bloqueo: no se puede modificar un pedido ya pagado
        paid, _ = self.get_payment_status(self.current_table)
        if paid:
            raise PermissionError("El pedido ya fue pagado y no puede modificarse.")
        self.table_orders[self.current_table].append(
            {"category": category, "dish": dish, "variant": variant, "price": price}
        )
        _write_audit("AGREGAR", self.current_table, f"{dish} ({variant}) Bs{price}")

    def remove_item_from_order(self, table_name: str, index: int) -> bool:
        # Bloqueo: no se puede modificar un pedido ya pagado
        paid, _ = self.get_payment_status(table_name)
        if paid:
            raise PermissionError("El pedido ya fue pagado y no puede modificarse.")
        if table_name in self.table_orders and 0 <= index < len(self.table_orders[table_name]):
            item = self.table_orders[table_name][index]
            _write_audit("ELIMINAR", table_name, f"{item['dish']} ({item['variant']})")
            del self.table_orders[table_name][index]
            return True
        return False

    def get_order_summary(self, table_name: str) -> Tuple[Dict[str, int], float]:
        order_summary = defaultdict(int)
        total = 0.0
        for item in self.table_orders.get(table_name, []):
            key = f"{item['dish']} ({item['variant']})"
            order_summary[key] += 1
            total += item["price"]
        return order_summary, total

    def delete_table(self, table_name: str):
        _write_audit("ELIMINAR_MESA", table_name)
        self.table_orders.pop(table_name, None)
        self.payment_status.pop(table_name, None)
        self.payment_details.pop(table_name, None)
        self.delivery_details.pop(table_name, None)
        self.order_numbers.pop(table_name, None)
        if self.current_table == table_name:
            self.current_table = None

    def rename_table(self, old_name: str, new_name: str) -> bool:
        if old_name not in self.table_orders:
            return False
        new_clean = " ".join((new_name or "").split()).strip()
        if not new_clean:
            return False
        if self._norm(old_name) == self._norm(new_clean):
            if old_name == new_clean:
                return True
            self.table_orders[new_clean] = self.table_orders.pop(old_name)
            for d in (self.payment_status, self.payment_details, self.delivery_details, self.order_numbers):
                if old_name in d:
                    d[new_clean] = d.pop(old_name)
            if self.current_table == old_name:
                self.current_table = new_clean
            return True
        if self.name_exists(new_clean):
            return False
        self.table_orders[new_clean] = self.table_orders.pop(old_name)
        for d in (self.payment_status, self.payment_details, self.delivery_details, self.order_numbers):
            if old_name in d:
                d[new_clean] = d.pop(old_name)
        if self.current_table == old_name:
            self.current_table = new_clean
        _write_audit("RENOMBRAR", old_name, f"nuevo={new_clean}")
        return True

    def get_all_tables(self) -> List[str]:
        return list(self.table_orders.keys())

    def clear_order(self, table_name: str):
        if table_name in self.table_orders:
            self.table_orders[table_name] = []

    def set_order_type(self, order_type: str):
        self.order_type = order_type

    # -----------------------------------------------
    #  Pagos
    # -----------------------------------------------
    def set_payment_status(
        self,
        table_name: str,
        paid: bool,
        method: str = "",
        amount_paid: float = 0,
        change: float = 0,
        cash_amount: float = 0,
        qr_amount: float = 0,
        change_method: str = "",
    ):
        self.payment_status[table_name] = (paid, method)
        if paid:
            _, total = self.get_order_summary(table_name)
            self.payment_details[table_name] = {
                "method": method,
                "cash_amount": cash_amount,
                "qr_amount": qr_amount,
                "amount_paid": amount_paid,
                "change": change,
                "change_method": change_method,
                "total": total,
            }
            num = self.order_numbers.get(table_name, "")
            _write_audit("PAGAR", table_name, f"num={num} metodo={method} total=Bs{total:.2f}")
            # Guardado automatico al Excel del dia
            try:
                self._append_to_daily_excel(table_name)
            except Exception as e:
                _write_audit("ERROR_EXCEL", table_name, str(e))

    def get_payment_status(self, table_name: str) -> Tuple[bool, str]:
        return self.payment_status.get(table_name, (False, ""))

    def get_payment_details(self, table_name: str) -> Optional[Dict]:
        return self.payment_details.get(table_name)

    # -----------------------------------------------
    #  Delivery
    # -----------------------------------------------
    def set_delivery_details(self, table_name: str, moto_cost: float, moto_payment_method: str):
        self.delivery_details[table_name] = {
            "moto_cost": moto_cost,
            "moto_payment_method": moto_payment_method,
        }

    def get_delivery_details(self, table_name: str) -> Optional[Dict]:
        return self.delivery_details.get(table_name)

    # -----------------------------------------------
    #  Excel diario (auto-guardado al pagar)
    # -----------------------------------------------
    def _append_to_daily_excel(self, table_name: str):
        """
        Agrega UNA fila al Excel del dia cuando se registra un pago.
        Crea el archivo y los headers si no existe todavia.
        """
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from collections import Counter

        data_dir = _get_data_dir()
        today = datetime.date.today().strftime("%Y-%m-%d")
        filepath = os.path.join(data_dir, f"pedidos_{today}.xlsx")

        headers_local = [
            "#", "Mesa/Cliente", "Platillos", "Total (Bs)",
            "Metodo de Pago", "Efectivo recibido (Bs)", "Monto QR (Bs)",
            "Cambio (Bs)", "Cambio dado en", "Pagado", "Hora"
        ]
        headers_delivery = [
            "#", "Mesa/Cliente", "Platillos", "Total (Bs)",
            "Metodo de Pago", "Efectivo recibido (Bs)", "Monto QR (Bs)",
            "Cambio (Bs)", "Cambio dado en",
            "Costo Moto (Bs)", "Pago Moto (metodo)", "Pagado", "Hora"
        ]

        def style_header(ws, headers):
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(1, col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1A5276")
                cell.alignment = Alignment(horizontal="center")

        # Cargar o crear workbook
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            ws_local    = wb["En el local"]    if "En el local"    in wb.sheetnames else wb.create_sheet("En el local")
            ws_delivery = wb["Para llevar"]   if "Para llevar"    in wb.sheetnames else wb.create_sheet("Para llevar")
        else:
            wb = Workbook()
            ws_local = wb.active
            ws_local.title = "En el local"
            ws_delivery = wb.create_sheet("Para llevar")
            style_header(ws_local, headers_local)
            style_header(ws_delivery, headers_delivery)

        items = self.table_orders.get(table_name, [])
        total = sum(i["price"] for i in items)
        details = self.get_payment_details(table_name) or {}
        delivery = self.get_delivery_details(table_name) or {}
        num = self.order_numbers.get(table_name, "")
        hora = datetime.datetime.now().strftime("%H:%M:%S")

        summary = Counter(f"{i['dish']} ({i['variant']})" for i in items)
        platillos_str = "; ".join(f"{k} x{v}" for k, v in summary.items())

        if delivery:
            ws_delivery.append([
                num, table_name, platillos_str, round(total, 2),
                details.get("method", ""),
                details.get("cash_amount") or "",
                details.get("qr_amount") or "",
                details.get("change") or "",
                details.get("change_method", ""),
                delivery.get("moto_cost", ""),
                delivery.get("moto_payment_method", ""),
                "Si", hora,
            ])
        else:
            ws_local.append([
                num, table_name, platillos_str, round(total, 2),
                details.get("method", ""),
                details.get("cash_amount") or "",
                details.get("qr_amount") or "",
                details.get("change") or "",
                details.get("change_method", ""),
                "Si", hora,
            ])

        # Auto-ajustar anchos
        for ws in [ws_local, ws_delivery]:
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        wb.save(filepath)

    # -----------------------------------------------
    #  Excel manual (boton Exportar)
    # -----------------------------------------------
    def save_all_to_excel(self, filepath: str = "pedidos.xlsx"):
        """
        Exporta TODOS los pedidos activos a un archivo elegido por el usuario.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from collections import Counter

        wb = Workbook()
        ws_local = wb.active
        ws_local.title = "En el local"
        headers_local = [
            "#", "Mesa/Cliente", "Platillos", "Total (Bs)",
            "Metodo de Pago", "Efectivo recibido (Bs)", "Monto QR (Bs)",
            "Cambio (Bs)", "Cambio dado en", "Pagado"
        ]
        ws_delivery = wb.create_sheet("Para llevar")
        headers_delivery = [
            "#", "Mesa/Cliente", "Platillos", "Total (Bs)",
            "Metodo de Pago", "Efectivo recibido (Bs)", "Monto QR (Bs)",
            "Cambio (Bs)", "Cambio dado en",
            "Costo Moto (Bs)", "Pago Moto (metodo)", "Pagado"
        ]

        def style_header_row(ws, headers):
            ws.append(headers)
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(1, col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1A5276")
                cell.alignment = Alignment(horizontal="center")

        style_header_row(ws_local, headers_local)
        style_header_row(ws_delivery, headers_delivery)

        for mesa, items in self.table_orders.items():
            if not items:
                continue
            total = sum(i["price"] for i in items)
            paid_status = self.get_payment_status(mesa)
            details = self.get_payment_details(mesa) or {}
            delivery = self.get_delivery_details(mesa) or {}
            num = self.order_numbers.get(mesa, "")
            summary = Counter(f"{i['dish']} ({i['variant']})" for i in items)
            platillos_str = "; ".join(f"{k} x{v}" for k, v in summary.items())

            if delivery:
                ws_delivery.append([
                    num, mesa, platillos_str, round(total, 2),
                    details.get("method", ""),
                    details.get("cash_amount") or "",
                    details.get("qr_amount") or "",
                    details.get("change") or "",
                    details.get("change_method", ""),
                    delivery.get("moto_cost", ""),
                    delivery.get("moto_payment_method", ""),
                    "Si" if paid_status[0] else "No",
                ])
            else:
                ws_local.append([
                    num, mesa, platillos_str, round(total, 2),
                    details.get("method", ""),
                    details.get("cash_amount") or "",
                    details.get("qr_amount") or "",
                    details.get("change") or "",
                    details.get("change_method", ""),
                    "Si" if paid_status[0] else "No",
                ])

        for ws in [ws_local, ws_delivery]:
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        wb.save(filepath)
